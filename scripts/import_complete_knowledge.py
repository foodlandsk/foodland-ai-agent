from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SECTION_SHEETS = (
    "Alternatives",
    "CrossSell",
    "Recipes",
    "Magazine",
    "FAQ",
    "Products_AI",
    "IntentMapping",
)


def clean_header(value: Any, index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"Unnamed: {index}"
    return str(value).strip()


def clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def sheet_to_records(workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_header(value, index) for index, value in enumerate(next(rows))]
    records: list[dict[str, Any]] = []

    for row in rows:
        record = {
            header: clean_value(value)
            for header, value in zip(headers, row)
            if not header.startswith("Unnamed:")
        }
        record = {key: value for key, value in record.items() if value != ""}
        if record:
            records.append(record)

    return records


def build_knowledge(xlsx_path: Path) -> dict[str, Any]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sections = {
        sheet_name: sheet_to_records(workbook, sheet_name)
        for sheet_name in SECTION_SHEETS
        if sheet_name in workbook.sheetnames
    }
    return {
        "version": "complete_v5_products_ai",
        "source": xlsx_path.name,
        "sections": sections,
        "counts": {section: len(records) for section, records in sections.items()},
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import complete Foodland knowledge workbook into data/knowledge.json.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--output", type=Path, default=Path("data/knowledge.json"))
    args = parser.parse_args()

    knowledge = build_knowledge(args.xlsx_path)
    args.output.write_text(
        json.dumps(knowledge, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(knowledge["counts"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
