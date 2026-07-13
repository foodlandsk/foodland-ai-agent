from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


LANGUAGE_COLUMNS = ("SK", "CZ", "AT", "EN", "PL", "HU", "VI")


def load_recipe_links(xlsx_path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(xlsx_path, read_only=False, data_only=True)
    sheet = workbook["Recepty Foodland"]
    headers = [cell.value for cell in sheet[1]]
    title_index = headers.index("Recept (SK názov)")
    language_indexes = {
        language: headers.index(language)
        for language in LANGUAGE_COLUMNS
        if language in headers
    }

    links_by_title: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2):
        title = row[title_index].value
        if not title:
            continue

        links: dict[str, str] = {}
        for language, index in language_indexes.items():
            cell = row[index]
            if cell.hyperlink and cell.hyperlink.target:
                links[f"{language}_url"] = cell.hyperlink.target

        if links:
            links_by_title[str(title).strip()] = links

    return links_by_title


def update_knowledge(knowledge_path: Path, links_by_title: dict[str, dict[str, str]]) -> tuple[int, int]:
    knowledge = json.loads(knowledge_path.read_text(encoding="utf-8"))
    recipes = knowledge.get("sections", {}).get("Recipes", [])
    matched = 0

    for record in recipes:
        title = str(record.get("Recept (SK názov)", "")).strip()
        links = links_by_title.get(title)
        if not links:
            continue
        record.update(links)
        matched += 1

    knowledge_path.write_text(
        json.dumps(knowledge, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return matched, len(recipes)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import recipe hyperlinks from Foodland recipe XLSX into knowledge.json.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--knowledge", type=Path, default=Path("data/knowledge.json"))
    args = parser.parse_args()

    links_by_title = load_recipe_links(args.xlsx_path)
    matched, total = update_knowledge(args.knowledge, links_by_title)
    print(f"Imported recipe links for {matched}/{total} recipes from {args.xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
