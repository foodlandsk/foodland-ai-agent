from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


LANGS = ["SK", "CZ", "AT", "EN", "PL", "HU", "VI"]

DEFAULT_FILES = {
    "Products_AI": "foodland_products_ai_tabulka.xlsx",
    "Products_Multilang": "Foodland Poradca/foodland_products_multilang.xlsx",
    "Product_Related_Live": "Foodland Poradca/foodland_products_multilang.xlsx",
    "Recipes": "foodland_recepty_jazykove_mutacie.xlsx",
    "Magazine": "foodland_magazin_clanky_jazykove_mutacie.xlsx",
    "CrossSell": "foodland_crosssell_tabulka.xlsx",
    "Alternatives": "foodland_alternativy_tabulka.xlsx",
    "FAQ": "foodland_faq_tabulka.xlsx",
    "IntentMapping": "foodland_intentmapping_tabulka.xlsx",
}

DEFAULT_SHEETS = {
    "Products_AI": ["Products_AI"],
    "Products_Multilang": ["Products"],
    "Product_Related_Live": ["Related Products"],
    "Recipes": ["Recepty Foodland", "Recipes", "Recepty"],
    "Magazine": ["Magazín články", "Magazin clanky", "Magazine"],
    "CrossSell": ["Cross-sell", "CrossSell"],
    "Alternatives": ["Alternativy", "Alternatives"],
    "FAQ": ["FAQ", "FAQs"],
    "IntentMapping": ["IntentMapping", "Intent Mapping"],
}

URL_COLUMNS = {
    "Products_AI": ["Produkt (URL)", "Cross-sell", "Alternatíva", "Súvisiaci recept"],
    "Recipes": LANGS,
    "Magazine": LANGS,
    "CrossSell": ["Produkt", "Cross-sell 1", "Cross-sell 2", "Cross-sell 3", "Cross-sell 4", "Cross-sell 5"],
    "Alternatives": ["Produkt", "Alternativa 1", "Alternativa 2", "Alternativa 3", "Alternativa 4", "Alternativa 5"],
    "FAQ": LANGS,
    "IntentMapping": LANGS,
}

CARRY_FORWARD_COLUMNS = {
    "Kategória",
    "Kategoria",
    "Kuchyňa",
    "Kuchyna",
    "Téma",
    "Tema",
    "Typ zámeru",
    "Typ zameru",
}


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_path = Path(args.output)

    sections: dict[str, list[dict[str, str]]] = {}
    source_files: dict[str, str] = {}

    if args.input_workbook:
        source_path = Path(args.input_workbook)
        workbook = load_workbook(source_path, data_only=False)
        for section in DEFAULT_FILES:
            records = read_section_from_workbook(section, workbook)
            sections[section] = records
            source_files[section] = str(source_path)
    else:
        for section in DEFAULT_FILES:
            source_path = input_dir / DEFAULT_FILES[section]
            if not source_path.exists():
                if args.allow_missing:
                    sections[section] = []
                    continue
                raise FileNotFoundError(f"Missing source file for {section}: {source_path}")

            records = read_section(section, source_path)
            sections[section] = records
            source_files[section] = str(source_path)

    payload = {
        "version": args.version,
        "source_dir": str(input_dir),
        "source_files": source_files,
        "sections": sections,
        "counts": {name: len(records) for name, records in sections.items()},
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(output_path)
    print(payload["counts"])


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Foodland Knowledge Excel files into data/knowledge.json.")
    parser.add_argument(
        "--input-dir",
        default="data/knowledge_sources",
        help="Directory with Foodland Excel source files.",
    )
    parser.add_argument(
        "--input-workbook",
        default="",
        help="Optional single workbook with all knowledge sheets. If set, --input-dir is ignored.",
    )
    parser.add_argument(
        "--output",
        default="data/knowledge.json",
        help="Output JSON path.",
    )
    parser.add_argument(
        "--version",
        default="Foodland_Knowledge_import_v1",
        help="Version label stored in the generated JSON.",
    )
    parser.add_argument(
        "--allow-missing",
        action="store_true",
        help="Keep missing sections empty instead of failing.",
    )
    return parser.parse_args()


def read_section(section: str, source_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(source_path, data_only=False)
    return read_section_from_workbook(section, workbook)


def read_section_from_workbook(section: str, workbook: Any) -> list[dict[str, str]]:
    if section == "Products_Multilang":
        return read_products_multilang(workbook)

    sheet = pick_sheet(workbook, DEFAULT_SHEETS[section])
    rows = list(sheet.iter_rows())
    if not rows:
        return []

    headers = [
        normalize_header(cell.value, index)
        for index, cell in enumerate(rows[0])
    ]
    records: list[dict[str, str]] = []
    carried: dict[str, str] = {}

    for excel_row in rows[1:]:
        record: dict[str, str] = {}
        has_value = False

        for header, cell in zip(headers, excel_row):
            value = cell_text(cell)
            has_value = has_value or bool(value)
            record[header] = value

            url = cell_url(cell)
            if should_add_url(section, header, url):
                record[f"{header}_url"] = url

        if not has_value:
            continue

        for key in CARRY_FORWARD_COLUMNS:
            if key not in record:
                continue
            if record[key]:
                carried[key] = record[key]
            elif carried.get(key):
                record[key] = carried[key]

        records.append(record)

    return records


def read_products_multilang(workbook: Any) -> list[dict[str, str]]:
    sheet = pick_sheet(workbook, DEFAULT_SHEETS["Products_Multilang"])
    rows = list(sheet.iter_rows())
    if len(rows) < 3:
        return []

    group_headers = [cell_text(cell) for cell in rows[0]]
    field_headers = [cell_text(cell) for cell in rows[1]]
    current_lang = ""
    headers: list[str] = []
    for group, field in zip(group_headers, field_headers):
        if group:
            current_lang = normalize_lang_group(group)
        if current_lang and field in {"Title", "URL", "Price", "Availability", "Image URL"}:
            headers.append(f"{current_lang}_{field}")
        else:
            headers.append(field or f"column_{len(headers) + 1}")

    records: list[dict[str, str]] = []
    for excel_row in rows[2:]:
        record: dict[str, str] = {}
        has_value = False
        for header, cell in zip(headers, excel_row):
            value = cell_text(cell)
            has_value = has_value or bool(value)
            record[header] = value

            url = cell_url(cell)
            if url and (header.endswith("_URL") or header.endswith("_Image URL")):
                record[header] = url

        if has_value:
            records.append(record)

    return records


def normalize_lang_group(value: str) -> str:
    text = str(value).strip().upper()
    if text.startswith("DE"):
        return "AT"
    if text in {"SK", "CZ", "AT", "EN", "PL", "HU", "VI"}:
        return text
    return text.replace(" ", "_").replace("(", "").replace(")", "")


def pick_sheet(workbook: Any, candidates: list[str]):
    for candidate in candidates:
        if candidate in workbook.sheetnames:
            return workbook[candidate]
    return workbook[workbook.sheetnames[0]]


def normalize_header(value: Any, index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text or f"column_{index + 1}"


def cell_text(cell: Any) -> str:
    value = "" if cell.value is None else str(cell.value).strip()
    if value in {"-", "–"}:
        return ""
    return value


def cell_url(cell: Any) -> str:
    if cell.hyperlink:
        url = cell.hyperlink.target or cell.hyperlink.location or ""
        return str(url).strip()
    value = cell_text(cell)
    if value.startswith(("http://", "https://")):
        return value
    return ""


def should_add_url(section: str, header: str, url: str) -> bool:
    return bool(url and header in URL_COLUMNS.get(section, []))


if __name__ == "__main__":
    main()
